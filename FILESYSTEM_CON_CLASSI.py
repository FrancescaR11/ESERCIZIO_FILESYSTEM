# -*- coding: utf-8 -*-
"""
Created on Tue Jan  5 18:14:47 2021

@author: gaiad,francescaronci
"""

import os 
import argparse 
import json  
from os.path import splitext 
import re  
import openpyxl 
import csv  
import pdfplumber 
from abc import ABC, abstractmethod
from Occurrence_Condition import Occurrence_Condition

'''

Uso argparse per la lettura del file di input e la scrittura del file di output.

'''

parser=argparse.ArgumentParser()

parser.add_argument("-i", "--input_data", help="Complete path to the file containing input data",
                    type=str, default='./dati/input.json')

parser.add_argument("-o", "--out_data", help="Complete path to the file containing output data", 
                    type=str, default='./results/output.txt')

args=parser.parse_args()


'''

Creo classe astratta sulle condizioni. Quetsa classe restituisce il path, il formato e le occorrenze delle parole presenti nel file di input.
Se il file è di un formato diverso da quelli considerati (pdf,txt,xlsx,csv), il programma stampa la stringa 'unknown file type'.

'''

class Conditions(ABC):
    """
    interface
    """
    @classmethod
    #@abstractmethod
    def path_condition(self):
        """
        abstract method
        """
        pass
    @classmethod
    #@abstractmethod
    def occurrence_condition(self):
        """
        abstract method
        """
        pass

    @staticmethod
    def create_instance(filename):  # Se ho file di testo indirizzo alle classi di controllo sul formato e sulle occorrenze (oppure alle future classi si controllo sul formato e sul contenuto delle immagini)
        suffix = splitext(filename)[1][1:].lower() # Estraggo il formato del file
        if (suffix == 'txt') | (suffix == 'csv') | (suffix == 'xlsx') | (suffix == 'pdf'): # Se il formato è tra quelli considerati, indirizzo alle classi PathFinder e OccurrenceFinder
            return PathFinder(filename), suffix , OccurrenceFinder(filename) # Restituisco il path, il formato e le occorrenze delle parole del file presenti nel file di input
        else:
            raise ValueError('unknown file type') 
            

'''

Creo la classe che prende in ingresso il file che stiamo analizzando ed il formato. Se il formato è presente nel file di input,
aggiungo il path del file alla lista dei file che soddisfano la prima condizione e chiamo la classe che verifica la condizione
sulle occorrenze.
Questa classe restitusice la lista dei path dei file che soddisfano la condizione sul formato e la lista dei file che soddisfano
sia la condizione sul formato che quella sulle occorrenze delle parole.

'''

class PathFinder(Conditions):

    def __init__(self, filename):
        
        self.filename = filename

    def path_condition(self,suffix):         
        
        for formato in dati[chiavi[1]]: #Scandisco i formati presenti nel file di input
            
            if suffix==formato: # Se il formato è presente tra i formati specificati nel file di input...
                  
             lista_cond_uno.append(file)  #... aggiungo il path del file a "lista_cond_uno", lista dei file che soddisfano la condizione sul formato
             
             lista_output=reader[2].occurrence_condition() # ...indirizzo alla classe che verifica la seconda condizione e salvo il path dei file che la verificano in "lista_output"
        
        # Restitusco la lista dei path dei file che soddisfano la condizione sul formato e la lista dei file che soddisfano sia la condizione sul formato che quella sulle occorrenze delle parole     
       
        return lista_cond_uno , lista_output 

'''

Creo la classe che chiama la funzione importata Occurence_Condition e resituisce in output la lista dei path dei file che soddisfano sia la condizione sul formato che quella sulle occorrenze delle parole

'''
 
class OccurrenceFinder(Conditions):

    def __init__(self, filename):
        
        self.filename = filename

    def occurrence_condition(self):     
        
        lista_output1=Occurrence_Condition(dati,lista_stringhe,chiavi,file,lista_output)
        
        return lista_output1


'''

Creo classe astratta che legge il formato. Questa classe prende in ingresso il file che sto analizzando e restituisce,
attraverso la chiamata alle classi derivate specifiche per ogni formato, la lista delle parole contenute nel file.

'''


class FormatReader(ABC):
    """
    interface
    """

    @abstractmethod
    def get_file_content(self):
        """
        abstract method
        """
        pass

    @staticmethod
    def create_instance(filename):
        suffix = splitext(filename)[1][1:].lower()
        if suffix == 'txt':
            return TXTReader(filename)
        elif suffix == 'csv':
            return CSVReader(filename)
        elif suffix == 'xlsx':
            return XLSXReader(filename)
        elif suffix == 'pdf':
            return PDFReader(filename)
        else:
            raise ValueError('unknown file type')

'''

Creo classe derivata per la lettura dei file txt. Questa classe prende in ingresso il file txt che sto analizzando e
restituisce la lista delle parole contenute nel file.

'''            

class TXTReader(FormatReader):

    def __init__(self, filename):
        
        self.filename = filename

    def get_file_content(self):
        
        f= open(file) 
        
        file_txt=(f.read()) # Salvo il testo contenuto nel file in una stringa
          
        lista_stringhe_txt=re.findall(r"[\w']+", file_txt.lower()) # Elimino la punteggiatura e divido la stringa (in minuscolo) ad ogni spazio
        
        return lista_stringhe_txt # Restituisco la lista delle parole contenute nel file
    

'''

Creo classe derivata per la lettura dei file csv. Questa classe prende in ingresso il file csv che sto analizzando e
restituisce la lista delle parole contenute nel file.

'''     

class CSVReader(FormatReader):

    def __init__(self, filename):
        
        self.filename = filename

    def get_file_content(self):
        
        with open(file) as filecsv:
            testo=' ' # Inizializzo una stringa vuota
            lettore = csv.reader(filecsv,delimiter=";") # Leggo il file csv
            for row in lettore: # Scandisco le righe dell'oggetto reader
                delimitat= " "
                testo=testo+(delimitat.join(row)).lower()+' ' # Aggiungo alla stringa 'testo' le righe del file csv (in minuscolo) separate dagli spazi
            lista_stringhe_csv=re.findall(r"[\w']+", testo) # Elimino la punteggiatura e divido le stringhe ad ogni spazio
        
        return lista_stringhe_csv # Restituisco la lista delle parole contenute nel file

'''

Creo classe derivata per la lettura dei file xlsx. Questa classe prende in ingresso il file xlsx che sto analizzando e
restituisce la lista delle parole contenute nel file.

'''         
    
class XLSXReader(FormatReader):

    def __init__(self, filename):
        self.filename = filename

    def get_file_content(self):
        
         stringhe=' ' # Inizializzo una stringa vuota
         excel_document = openpyxl.load_workbook(file) # Apro file excel
         nomi_fogli=(excel_document.sheetnames) # Salvo il nome dei fogli in una lista
   
         for nome in nomi_fogli: # Scandisco tutti i fogli del file excel
          sheet = excel_document[nome]

          for row in sheet.iter_rows(): # Scandisco le righe dei fogli excel
    
            for cell in row: # Scandisco le celle per riga
     
              if (cell.value is None)==False: # Se il valore della cella non è None...
       
                  stringhe=stringhe+((cell.value).lower())+' ' # ... aggiungo il contenuto della cella (in minuscolo) alla stringa                   
                  lista_stringhe_xlsx=re.findall(r"[\w']+", stringhe)  # Elimino la punteggiatura e divido le stringhe ad ogni spazio
        
         return lista_stringhe_xlsx # Restituisco la lista delle parole contenute nel file

'''

Creo classe derivata per la lettura dei file pdf. Questa classe prende in ingresso il file pdf che sto analizzando e
restituisce la lista delle parole contenute nel file.

'''     
    

class PDFReader(FormatReader):

    def __init__(self, filename):
        self.filename = filename

    def get_file_content(self):
        
        with pdfplumber.open (file) as pdf: # Apro il file pdf
           pages = pdf.pages  # Salvo le pagine del file pdf in una lista
           stringhe_pdf=' ' # Inizializzo una stringa vuota
           for page in pages: # Scandisco ogni pagina contenuta nella lista di pagine
              stringhe_pdf=stringhe_pdf+ (page.extract_text ()).lower() + ' ' # Aggiungo il contenuto della pagina (in minuscolo) alla stringa 
           lista_stringhe_pdf=re.findall(r"[\w']+", stringhe_pdf)  # Elimino la punteggiatura e divido le stringhe ad ogni spazio
           
           
        return lista_stringhe_pdf # Restituisco la lista delle parole contenute nel file
    


                            ########## main #########    
        
# Apertura file input
            
with open(args.input_data) as json_file:
   dati=json.load(json_file)


#Salvo nella lista "chaivi" il nome delle chiavi del dizionario di input

chiavi=list(dati.keys()) 

lista_path=[] # Inizializzo lista che conterrà i path dei file trovati

for path in dati[chiavi[0]]: # Scandisco tutti i path contenuti nel file input
 for root,dirs, files in os.walk(path, topdown=False): #Scandisco sottocartelle (dirs) e files
   for name in files: # Scandisco i file presenti nelle sottocartelle
       if name!= ".DS_Store":
        lista_path.append(os.path.join(root, name)) # Aggiungo il path del file 'name' alla lista dei path dei file trovati
        


lista_cond_uno=[] # Inizializzo la lista che conterrà i path dei file che soddisfano la condizione sul formato    

lista_output=[] # Inizializzo la lista che conterrà i file che soddisfano sia la condizione sul formato che quella sulle occorrenze delle parole
   
for file in lista_path:    # Scandisco tutti i file trovati nei path passati con il file di input    
                   
    reader = FormatReader.create_instance(file) # Chiamo classe FormatReader per leggere il formato del file
    
    lista_stringhe = reader.get_file_content() # Estraggo la lista di stringhe con la funzione get_file_content()
    
    reader = Conditions.create_instance(file) # Chiamo classe Conditions per estrarre le liste dei file che rispettano le condizioni
    
    #reader[0] contiene il path del file
    #reader[1] contiene il formato del file
    
    lista_cond_uno,lista_output= reader[0].path_condition(reader[1]) # Estraggo le liste dei file che rispettano le condizioni con la funzione path_condition contenuta in PathFinder


# Salvo su un file in formato txt la lista dei file che rispettano le condizioni

with open(args.out_data, "w") as output:
    
    output.write(str(lista_output))  







