#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Jan  9 13:33:08 2021

@author: francescaronci,gaiad
"""


import os 
import argparse 
import json  
from os.path import splitext 
import re  
import openpyxl 
import csv  
import pdfplumber 
from abc import ABC, abstractmethod
from Occurrence_Condition import Occurrence_Condition

'''

Uso argparse per la lettura del file di input e la scrittura del file di output.

'''

parser=argparse.ArgumentParser()

parser.add_argument("-i", "--input_data", help="Complete path to the file containing input data",
                    type=str, default='./dati/input.json')

parser.add_argument("-o", "--out_data", help="Complete path to the file containing output data", 
                    type=str, default='./results/output.txt')

args=parser.parse_args()


'''

Creo la classe che prende in ingresso il file che stiamo analizzando. 
Se il formato è presente nel file di input, chiamo la classe che verifica la 
seconda condizione.


'''

class PathChecker():
    
    def __init__(self, filename):
        
       self.filename=filename

        
    def checker(filename) :
       
        suffix = splitext(filename)[1][1:].lower()
        

            
        if suffix in dati['filetypelist']: # Se il formato del file è presente tra i formati del file di input...
               
               # Se il file contiene una stringa chiamo la classe che conta le occorrenze
               
               if (suffix == 'txt') | (suffix == 'csv') | (suffix == 'xlsx') | (suffix == 'pdf'): 
            
                   return OccurrenceChecker(filename)
               
                # Se il file è un immagine chiamo la classe che opera sulle immagini
               
               elif suffix=='jpeg':
                   
                   return ImageChecker(filename)
        else:
                
                return 'failure'
            
            
            
            
'''

Creo classe astratta sulle condizioni. Questa classe rimanda alla classe PathChecker,
indipendentemente dal formato.
Questa classe contiene il metodo astratto checker(), che viene implementato diversamente
in base alla condizione.

'''

class Condition(ABC):
    """
    interface
    """
    def __init__(self,filename):
       self.filename=filename

    @staticmethod
    
    def create_instance(filename):  

        return PathChecker.checker(filename)
            
    @abstractmethod
    def checker(self) :
        """
        abstract method
        """
        pass
        
    

'''

Creo la classe derivata che chiama la funzione importata Occurence_Condition e resituisce in output
la lista dei path dei file che soddisfano sia la condizione sul formato che quella sulle occorrenze delle parole.

'''
 
class OccurrenceChecker(Condition):
    
    def __init__(self,filename):
        
      self.filename=filename 

        
    
    def checker(self,filename):   
        
       parametro=FormatReader.create_instance(filename).get_file_content()
       
       if (Occurrence_Condition(dati,parametro))== True:
        
        return True
       
       else:
       
        return False
    
    

'''
Creo la classe derivata che verifica la seconda condizione sulle immagini.

'''   

class ImageChecker(Condition):
   
    def __init__(self, filename):
        
       super().__init__(filename)   
    
    
    def checker(self):
        pass
 
     


'''

Creo classe astratta che legge il formato. Questa classe prende in ingresso il file che sto analizzando e restituisce,
attraverso la chiamata alle classi derivate specifiche per ogni formato, il contenuto del file.
Nel caso di file di testo, restituisce la lista delle parole.

'''


class FormatReader(ABC):
    """
    interface
    """
    def __init__(self,filename):
        self.filename=filename

    @abstractmethod
    def get_file_content(self):
        """
        abstract method
        """
        pass

    @staticmethod
    def create_instance(filename):
        
        suffix = splitext(filename)[1][1:].lower()
 
        
        if suffix == 'txt':
        
                return TXTReader(filename)
        
        elif suffix == 'csv':
            
                 return CSVReader(filename)
        
        elif suffix == 'xlsx':
            
                 return XLSXReader(filename)
        
        elif suffix == 'pdf':
            
                return PDFReader(filename)
        
        else:
            
             raise ValueError('unknown file type')

'''

Creo classe derivata per la lettura dei file txt. Questa classe prende in ingresso il file txt che sto analizzando e
restituisce la lista delle parole contenute nel file.

'''            

class TXTReader(FormatReader):

    def __init__(self, filename):
        
       super().__init__(filename)

    
    def get_file_content(self):
        
        f= open(file) 
        
        file_txt=(f.read()) # Salvo il testo contenuto nel file in una stringa
          
        lista_stringhe_txt=re.findall(r"[\w']+", file_txt.lower()) # Elimino la punteggiatura e divido la stringa (in minuscolo) ad ogni spazio
        
        return lista_stringhe_txt # Restituisco la lista delle parole contenute nel file
    

'''

Creo classe derivata per la lettura dei file csv. Questa classe prende in ingresso il file csv che sto analizzando e
restituisce la lista delle parole contenute nel file.

'''     

class CSVReader(FormatReader):

    def __init__(self, filename):
        
        super().__init__(filename)

   
    def get_file_content(self):
        
        with open(file) as filecsv:
            
            testo=' ' # Inizializzo una stringa vuota
            
            lettore = csv.reader(filecsv,delimiter=";") # Leggo il file csv
            
            for row in lettore: # Scandisco le righe dell'oggetto reader
                
                delimitat= " "
                
                # Aggiungo alla stringa 'testo' le righe del file csv (in minuscolo) separate dagli spazi
                
                testo=testo+(delimitat.join(row)).lower()+' ' 
           
            # Elimino la punteggiatura e divido le stringhe ad ogni spazio
            
            lista_stringhe_csv=re.findall(r"[\w']+", testo) 
        
        return lista_stringhe_csv # Restituisco la lista delle parole contenute nel file

'''

Creo classe derivata per la lettura dei file xlsx. Questa classe prende in ingresso il file xlsx che sto analizzando e
restituisce la lista delle parole contenute nel file.

'''         
    
class XLSXReader(FormatReader):

    def __init__(self, filename):
        super().__init__(filename)

   
    def get_file_content(self):
        
         stringhe=' ' # Inizializzo una stringa vuota
         
         excel_document = openpyxl.load_workbook(file) # Apro file excel
        
         nomi_fogli=(excel_document.sheetnames) # Salvo il nome dei fogli in una lista
   
         for nome in nomi_fogli: # Scandisco tutti i fogli del file excel
          
          sheet = excel_document[nome]

          for row in sheet.iter_rows(): # Scandisco le righe dei fogli excel
    
            for cell in row: # Scandisco le celle per riga
     
              if (cell.value is None)==False: # Se il valore della cella non è None...
       
                  stringhe=stringhe+((cell.value).lower())+' ' # ... aggiungo il contenuto della cella (in minuscolo) alla stringa                   
                 
                  lista_stringhe_xlsx=re.findall(r"[\w']+", stringhe)  # Elimino la punteggiatura e divido le stringhe ad ogni spazio
        
         return lista_stringhe_xlsx # Restituisco la lista delle parole contenute nel file

'''

Creo classe derivata per la lettura dei file pdf. Questa classe prende in ingresso il file pdf che sto analizzando e
restituisce la lista delle parole contenute nel file.

'''     
    

class PDFReader(FormatReader):

    def __init__(self, filename):
        super().__init__(filename)

   
    def get_file_content(self):
        
        with pdfplumber.open (file) as pdf: # Apro il file pdf
           
           pages = pdf.pages  # Salvo le pagine del file pdf in una lista
           
           stringhe_pdf=' ' # Inizializzo una stringa vuota
           
           for page in pages: # Scandisco ogni pagina contenuta nella lista di pagine
                 
                 # Aggiungo il contenuto della pagina (in minuscolo) alla stringa
                 
                 stringhe_pdf=stringhe_pdf+ (page.extract_text ()).lower() + ' '  
           
           lista_stringhe_pdf=re.findall(r"[\w']+", stringhe_pdf)  # Elimino la punteggiatura e divido le stringhe ad ogni spazio
           
           
        return lista_stringhe_pdf # Restituisco la lista delle parole contenute nel file
    


                            ########## main #########    
        
# Apertura file input
            
with open(args.input_data) as json_file:
   
    dati=json.load(json_file)



lista_path=[] # Inizializzo lista che conterrà i path dei file trovati

for path in dati['dirlist']: # Scandisco tutti i path contenuti nel file input
 
  for root,dirs, files in os.walk(path, topdown=False): #Scandisco sottocartelle (dirs) e files
   
    for name in files: # Scandisco i file presenti nelle sottocartelle
       
       if name!= ".DS_Store":
        
             lista_path.append(os.path.join(root, name)) # Aggiungo il path del file 'name' alla lista dei path dei file trovati
        

   

# Inizializzo la lista che conterrà i file che soddisfano sia la condizione sul formato che quella sulle occorrenze delle parole

lista_output=[] 


for file in lista_path: # Scandisco tutti i file trovati nei path passati con il file di input
    
    condizione_verificata=True
    condizioni_finite=False 
 
    condition=Condition.create_instance(file)
    

     
    if condition is not ['failure']:
        
        
         
         while (condizione_verificata) & (not(condizioni_finite)):
         
          try:
             condizione_verificata= condition.checker(file)
             
             condition=condizione_verificata
             
  
             
         
          except AttributeError:
              
             condizioni_finite=True
             
             print('Condizioni da verificare terminate')
             
        
         if condizione_verificata==True :
             
             lista_output.append(file)
        
         else :
             print ('Non tutte le condizioni sono verificate')
   
    else :
         print ('Il formato del file non è tra quelli cercati')
        
             



# Salvo su un file in formato txt la lista dei file che rispettano le condizioni

with open(args.out_data, "w") as output:
    
    output.write(str(lista_output))  
