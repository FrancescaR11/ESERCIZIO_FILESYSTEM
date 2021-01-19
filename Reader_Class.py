#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jan 19 13:30:24 2021

@author: francescaronci
"""
import os  
from os.path import splitext 
import re  
import openpyxl 
import csv  
import pdfplumber 
from abc import ABC, abstractmethod
import argparse 



parser=argparse.ArgumentParser()

parser.add_argument("-o_2", "--out_data_2", help="Complete path to the file containing output of the images detector", 
                    type=str, default='./results/')

args=parser.parse_args()

'''

Creo classe astratta 'FormatReader' che legge il formato. Questa classe prende in ingresso il file 
che sto analizzando e restituisce, attraverso la chiamata alle classi derivate specifiche per ogni formato,
il contenuto del file. Nel caso di file di testo, restituisce la lista delle parole. Nel caso di file immagine,
restituisce la lista degli oggetti presenti nell'immagine.

'''


class FormatReader(ABC):
    
    """
    interface
    """
   
    def __init__(self,filename):  
        
      self.filename=filename
      super().__init__()

    @abstractmethod
    def size_extractor(self):

      file_size=os.path.getsize (self.filename)
      
      return file_size
  
    @abstractmethod
    def time_extractor(self):
        
        file_time=os.path.getctime(self.filename)
        
        return file_time

    @abstractmethod
    def get_file_content(self):
        """
        abstract method
        """
        pass

    @staticmethod
    
    #Il metodo 'create_instance()' rimanda alla classe derivata specifica per ogni formato.
    def create_instance(filename):
        
        suffix = splitext(filename)[1][1:].lower()
 
        
        if suffix == 'txt':
        
                return TXTReader(filename)
        
        elif suffix == 'csv':
            
                 return CSVReader(filename)
        
        elif suffix == 'xlsx':
            
                 return XLSXReader(filename)
        
        elif suffix == 'pdf':
            
                return PDFReader(filename)
            
            
        elif suffix=='jpeg':
            
                return JPEGReader(filename)    
        
        else:
            
             raise ValueError('unknown file type')


"""
Creo classe derivata per la lettura dei file jpeg. Questa classe prende in ingresso il file jpeg
che sto analizzando e restituisce la lista degli oggetti contenuti nel file immagine.

"""

class JPEGReader(FormatReader):
        
    def size_extractor(self):
        
        file_size=super().size_extractor()
        
        return file_size
  
    def time_extractor(self):
        
        file_time=super().size_extractor()
        
        return file_time   

    
    def get_file_content(self,detector):
        
        found_objects=[]  #Inizializzo la lista degli oggetti trovati nell'immagine
        #detector = ObjectDetection()

        model_path = "./models/yolo-tiny.h5"
        # #output_path = args.out_dir
        
        detector.setModelTypeAsTinyYOLOv3()
        detector.setModelPath(model_path)
        detector.loadModel()
        
        #creo la lista detection contenente gli oggetti rilevati nell'immagine con una probabilità superiore al 30%. 
        
        detection = detector.detectObjectsFromImage(input_image=self.filename, output_image_path=args.out_data_2+os.path.basename(self.filename),minimum_percentage_probability=30)

        
        for eachItem in detection:
            
            #Appendo alla lista 'found_objects' i nomi degli oggetti trovati
            found_objects.append(eachItem["name"]) 
            
        return found_objects  # Restituisco la lista degli oggetti contenuti nel file immagine
    



'''

Creo classe derivata per la lettura dei file txt. Questa classe prende in ingresso il file txt che sto analizzando e
restituisce la lista delle parole contenute nel file.

'''            

class TXTReader(FormatReader):

    def size_extractor(self):
        
        file_size=super().size_extractor()
        
        return file_size
  
    def time_extractor(self):
        
        file_time=super().size_extractor()
        
        return file_time 
    
   
    def get_file_content(self):
        
        f= open(self.filename) 
        
        file_txt=(f.read()) # Salvo il testo contenuto nel file in una stringa
          
        lista_stringhe_txt=re.findall(r"[\w']+", file_txt.lower()) # Elimino la punteggiatura e divido la stringa (in minuscolo) ad ogni spazio
        
        return lista_stringhe_txt # Restituisco la lista delle parole contenute nel file
    

'''

Creo classe derivata per la lettura dei file csv. Questa classe prende in ingresso il file csv che sto analizzando e
restituisce la lista delle parole contenute nel file.

'''     

class CSVReader(FormatReader):

    def size_extractor(self):
        
        file_size=super().size_extractor()
        
        return file_size
  
    def time_extractor(self):
        
        file_time=super().size_extractor()
        
        return file_time 
  
    def get_file_content(self):
        
        with open(self.filename) as filecsv:
            
            testo=' ' # Inizializzo una stringa vuota
            
            lettore = csv.reader(filecsv,delimiter=";") # Leggo il file csv
            
            for row in lettore: # Scandisco le righe dell'oggetto reader
                
                delimitat= " "
                
                # Aggiungo alla stringa 'testo' le righe del file csv (in minuscolo) separate dagli spazi
                
                testo=testo+(delimitat.join(row)).lower()+' ' 
           
            # Elimino la punteggiatura e divido le stringhe ad ogni spazio
            
            lista_stringhe_csv=re.findall(r"[\w']+", testo) 
        
        return lista_stringhe_csv # Restituisco la lista delle parole contenute nel file

'''

Creo classe derivata per la lettura dei file xlsx. Questa classe prende in ingresso il file xlsx che sto analizzando e
restituisce la lista delle parole contenute nel file.

'''         
    
class XLSXReader(FormatReader):

    def size_extractor(self):
        
        file_size=super().size_extractor()
        
        return file_size
  
    def time_extractor(self):
        
        file_time=super().size_extractor()
        
        return file_time 
    
    def get_file_content(self):
        
         stringhe=' ' # Inizializzo una stringa vuota
         
         excel_document = openpyxl.load_workbook(self.filename) # Apro file excel
        
         nomi_fogli=(excel_document.sheetnames) # Salvo il nome dei fogli in una lista
   
         for nome in nomi_fogli: # Scandisco tutti i fogli del file excel
          
          sheet = excel_document[nome]

          for row in sheet.iter_rows(): # Scandisco le righe dei fogli excel
    
            for cell in row: # Scandisco le celle per riga
     
              if (cell.value is None)==False: # Se il valore della cella non è None...
       
                  stringhe=stringhe+((cell.value).lower())+' ' # ... aggiungo il contenuto della cella (in minuscolo) alla stringa                   
                 
                  lista_stringhe_xlsx=re.findall(r"[\w']+", stringhe)  # Elimino la punteggiatura e divido le stringhe ad ogni spazio
        
         return lista_stringhe_xlsx # Restituisco la lista delle parole contenute nel file

'''

Creo classe derivata per la lettura dei file pdf. Questa classe prende in ingresso il file pdf che sto analizzando e
restituisce la lista delle parole contenute nel file.

'''     
    

class PDFReader(FormatReader):
   
    def size_extractor(self):
        
        file_size=super().size_extractor()
        
        return file_size
  
    def time_extractor(self):
        
        file_time=super().size_extractor()
        
        return file_time 
   
    def get_file_content(self):
        
        with pdfplumber.open (self.filename) as pdf: # Apro il file pdf
           
           pages = pdf.pages  # Salvo le pagine del file pdf in una lista
           
           stringhe_pdf=' ' # Inizializzo una stringa vuota
           
           for page in pages: # Scandisco ogni pagina contenuta nella lista di pagine
                 
                 # Aggiungo il contenuto della pagina (in minuscolo) alla stringa
                 
                 stringhe_pdf=stringhe_pdf+ (page.extract_text ()).lower() + ' '  
           
           lista_stringhe_pdf=re.findall(r"[\w']+", stringhe_pdf)  # Elimino la punteggiatura e divido le stringhe ad ogni spazio
           
           
        return lista_stringhe_pdf # Restituisco la lista delle parole contenute nel file