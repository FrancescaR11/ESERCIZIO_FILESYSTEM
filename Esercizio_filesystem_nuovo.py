# -*- coding: utf-8 -*-
"""
Created on Tue Jan 19 11:52:14 2021

@author: gaiad
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Jan 14 12:19:09 2021

@author: francescaronci,gaiad
"""

import os 
import argparse 
import json  
from os.path import splitext 
import re  
import openpyxl 
import csv  
import pdfplumber 
from abc import ABC, abstractmethod
import datetime,time
from imageai.Detection import ObjectDetection


'''

Uso argparse per la lettura dei file di input e la scrittura del file di output.
Il programma riceve in ingresso:
    
- 'input.json' --> il file json contentente la porzione del filesystem su cui effettuare la ricerca e
  una lista delle condizioni che devono essere verificate.
  
- 'objects.txt' --> una file di testo contenente la lista degli oggetti individuabili dall'algoritmo

Il programma restituisce in output un file txt contenente la lista di file che soddisfano le condizioni richieste.

'''

parser=argparse.ArgumentParser()

parser.add_argument("-i", "--input_data", help="Complete path to the file containing input data",
                    type=str, default='./dati/input.json')

parser.add_argument("-o", "--out_data", help="Complete path to the file containing output data", 
                    type=str, default='./results/output.txt')

parser.add_argument("-i_2", "--input_data_2", help="Complete path to the file containing output data", 
                    type=str, default='./dati/objects.txt')

parser.add_argument("-o_2", "--out_data_2", help="Complete path to the file containing output of the images detector", 
                    type=str, default='./results/')

args=parser.parse_args()


'''

Creo la classe 'PathChecker' che prende in ingresso il file che stiamo analizzando e il dizionario, 'condition_list'
contenente l'informazione del file json di input. 
Se il formato è presente nel file di input restituisce True altrimenti restituisce False


'''

class PathChecker():
    
    def __init__(self,filename,condition_list):
        
     self.filename=filename
     self.condition_list=condition_list
        
    def function(self) :
       
        suffix = splitext(self.filename)[1][1:].lower()  #Estraggo il formato del file
  
        #Se il formato del file è presente tra i formati del file di input restituisco True
        if suffix in self.condition_list['filetypelist']: 
               
               return True
        else:
                
               return False
            
                       
            
'''

Creo classe astratta sulle condizioni. Questa classe rimanda, tramite il metodo 'create_instance()' 
alla classe PathChecker, indipendentemente dal formato.
Questa classe contiene il metodo astratto checker(), che viene implementato diversamente
in base alla condizione. Tale metodo restituisce sempre 'True' se la condizione è soddisfatta
e 'False' altrimenti.

'''

class Condition(ABC):
    """
    interface
    """
    def __init__(self):
     
     pass
            
    @abstractmethod
  
    def checker(self) :
        
        """
        abstract method
        """
        pass
        
    

'''

Creo la classe derivata 'OccurrenceChecker' che verifica, tramite il metodo 'checker()',
per ogni coppia parola-occorrenza, che la condizione sia verificata.

'''
 
class OccurrenceChecker(Condition):
    
    def __init__(self,parola,occorrenza):
       
     self.parola= parola
     self.occorrenza=occorrenza
     #super().__init__(condition_list)
    

        
    
    def checker(self,file_object):
     
     #Se il 'file_object' creato fa riferimento ad un formato immagine,
     #restituisce 'True' e salta la condizione relativa al file di testo.
     
     if   type(file_object)== JPEGReader:
         
         return True 
     
     else:
         
         #Il metodo 'get_file_content' restituisce la lista di stringhe presente nel file di testo
     
         lista_parole=file_object.get_file_content()
 
         ripetizioni=0  #Inizializzo il numero di ripetizioni della parola

         for stringa in lista_parole: #Confronto ogni stringa di 'lista_parole' con la parola cercata
             
           if stringa==self.parola: #Se è presente nel file..
                
              ripetizioni+=1 #..incremento
         
         if ripetizioni >= self.occorrenza: #Se la stringa è presente almeno tante volte quanto richiesto dalla condizione..
              
               return True 
         else :
              
               return False

            

'''
Creo la classe derivata 'SizeChecker' che verifica, tramite il metodo 'checker()',
che le condizioni sulla dimensione del file siano verificate.

'''


class SizeChecker(Condition):
    
    
    def __init__(self,condition_list,min_value,max_value):
        
        
        self.condition_list=condition_list
        #super().__init__(condition_list)
        self.min_value=min_value
        self.max_value=max_value
       
    
    def checker(self,file_object):
        
          if self.min_value=='':
              
              self.min_value=0
            
          if (self.max_value=='') | (self.max_value=='infinito'):
              
              self.max_value=float('inf')
              
                
          #Verifico la condizione sulla dimensione
          if  self.min_value < file_object.size_extractor() < self.max_value :
          
              return True
          
          else:
             
              return False

            
                       
'''
Creo la classe derivata 'TimeChecker' che verifica, tramite il metodo 'checker()',
che le condizioni sulla data di creazione del file siano verificate.

'''

class TimeChecker(Condition):
    
    def __init__(self,condition_list,min_value,max_value):
        
        
        #super().__init__(condition_list)
        self.condition_list=condition_list
        self.min_value=min_value
        self.max_value=max_value
        
        
    def checker(self,file_object):   
        
            
            #Se il minimo non è specificato lo sostituisco con 0
            if self.min_value== '':
                
                time_min=0
                
            #Se invece il minimo è specificato    
            else :
                
                #Estraggo la data di creazione del file e la converto il formato timestamp
                time_min=time.mktime(datetime.datetime.strptime(self.min_value, "%Y-%m-%d").timetuple())  
                
            if self.max_value=='':
                
                time_max=float('inf')
                
            else :
                #Estraggo la data di creazione del file e la converto il formato timestamp
                time_max=time.mktime(datetime.datetime.strptime(self.max_value, "%Y-%m-%d").timetuple())
                
            #Verifico la condizione sulla data di creazione
            if  (time_min) < file_object.time_extractor() < (time_max):
            
                return True
            
            else:
                
                return False


'''
Creo la classe derivata 'ImageChecker' che verifica, tramite il metodo 'checker()',
per ogni coppia obj-occurrence, che la condizione sia verificata.

'''   

class ImageChecker(Condition):
   
    def __init__(self, condition_list, obj, occurrence, detector):
       
        
       self.condition_list=condition_list
       self.obj=obj
       self.occurrence=occurrence
       self.detector=detector
       
    
    
    def checker(self,file_object):
        
        #Se il 'file_object' creato fa riferimento ad un formato testuale,
        #restituisce 'True' e salta la condizione relativa al file immagine.
        if type(file_object)!= JPEGReader:
            
            return True
        
        else:
            
            #Il metodo 'get_file_content' restituisce la lista di oggetti presente nel file immagine
            found_objects=file_object.get_file_content(self.detector)
        
            ripetizioni=0  #Inizializzo il numero di ripetizioni dell'oggetto 

            for found_obj in found_objects: #Confronto ogni oggetto di 'found_objects' con l'oggetto cercato
             
              if found_obj==self.obj: #Se l'oggetto è presente..
                
               ripetizioni+=1  #..incremento
         
            if ripetizioni >= self.occurrence: #Se l'oggetto' è presente almeno tante volte quanto richiesto dalla condizione..
              
               return True
            else :
              
               return False
        
                     
            
    


'''

Creo classe astratta 'FormatReader' che legge il formato. Questa classe prende in ingresso il file 
che sto analizzando e restituisce, attraverso la chiamata alle classi derivate specifiche per ogni formato,
il contenuto del file. Nel caso di file di testo, restituisce la lista delle parole. Nel caso di file immagine,
restituisce la lista degli oggetti presenti nell'immagine.

'''


class FormatReader(ABC):
    
    """
    interface
    """
   
    def __init__(self,filename):  
        
      self.filename=filename
      super().__init__()

    @abstractmethod
    def size_extractor(self):

      file_size=os.path.getsize (self.filename)
      
      return file_size
  
    @abstractmethod
    def time_extractor(self):
        
        file_time=os.path.getctime(self.filename)
        
        return file_time

    @abstractmethod
    def get_file_content(self):
        """
        abstract method
        """
        pass

    @staticmethod
    
    #Il metodo 'create_instance()' rimanda alla classe derivata specifica per ogni formato.
    def create_instance(filename):
        
        suffix = splitext(filename)[1][1:].lower()
 
        
        if suffix == 'txt':
        
                return TXTReader(filename)
        
        elif suffix == 'csv':
            
                 return CSVReader(filename)
        
        elif suffix == 'xlsx':
            
                 return XLSXReader(filename)
        
        elif suffix == 'pdf':
            
                return PDFReader(filename)
            
            
        elif suffix=='jpeg':
            
                return JPEGReader(filename)    
        
        else:
            
             raise ValueError('unknown file type')


"""
Creo classe derivata per la lettura dei file jpeg. Questa classe prende in ingresso il file jpeg
che sto analizzando e restituisce la lista degli oggetti contenuti nel file immagine.

"""

class JPEGReader(FormatReader):
        
    def size_extractor(self):
        
        file_size=super().size_extractor()
        
        return file_size
  
    def time_extractor(self):
        
        file_time=super().size_extractor()
        
        return file_time   

    
    def get_file_content(self,detector):
        
        found_objects=[]  #Inizializzo la lista degli oggetti trovati nell'immagine
        #detector = ObjectDetection()

        model_path = "./models/yolo-tiny.h5"
        # #output_path = args.out_dir
        
        detector.setModelTypeAsTinyYOLOv3()
        detector.setModelPath(model_path)
        detector.loadModel()
        
        #creo la lista detection contenente gli oggetti rilevati nell'immagine con una probabilità superiore al 30%. 
        
        detection = detector.detectObjectsFromImage(input_image=self.filename, output_image_path=args.out_data_2+os.path.basename(self.filename),minimum_percentage_probability=30)

        
        for eachItem in detection:
            
            #Appendo alla lista 'found_objects' i nomi degli oggetti trovati
            found_objects.append(eachItem["name"]) 
            
        return found_objects  # Restituisco la lista degli oggetti contenuti nel file immagine
    



'''

Creo classe derivata per la lettura dei file txt. Questa classe prende in ingresso il file txt che sto analizzando e
restituisce la lista delle parole contenute nel file.

'''            

class TXTReader(FormatReader):

    def size_extractor(self):
        
        file_size=super().size_extractor()
        
        return file_size
  
    def time_extractor(self):
        
        file_time=super().size_extractor()
        
        return file_time 
    
   
    def get_file_content(self):
        
        f= open(self.filename) 
        
        file_txt=(f.read()) # Salvo il testo contenuto nel file in una stringa
          
        lista_stringhe_txt=re.findall(r"[\w']+", file_txt.lower()) # Elimino la punteggiatura e divido la stringa (in minuscolo) ad ogni spazio
        
        return lista_stringhe_txt # Restituisco la lista delle parole contenute nel file
    

'''

Creo classe derivata per la lettura dei file csv. Questa classe prende in ingresso il file csv che sto analizzando e
restituisce la lista delle parole contenute nel file.

'''     

class CSVReader(FormatReader):

    def size_extractor(self):
        
        file_size=super().size_extractor()
        
        return file_size
  
    def time_extractor(self):
        
        file_time=super().size_extractor()
        
        return file_time 
  
    def get_file_content(self):
        
        with open(self.filename) as filecsv:
            
            testo=' ' # Inizializzo una stringa vuota
            
            lettore = csv.reader(filecsv,delimiter=";") # Leggo il file csv
            
            for row in lettore: # Scandisco le righe dell'oggetto reader
                
                delimitat= " "
                
                # Aggiungo alla stringa 'testo' le righe del file csv (in minuscolo) separate dagli spazi
                
                testo=testo+(delimitat.join(row)).lower()+' ' 
           
            # Elimino la punteggiatura e divido le stringhe ad ogni spazio
            
            lista_stringhe_csv=re.findall(r"[\w']+", testo) 
        
        return lista_stringhe_csv # Restituisco la lista delle parole contenute nel file

'''

Creo classe derivata per la lettura dei file xlsx. Questa classe prende in ingresso il file xlsx che sto analizzando e
restituisce la lista delle parole contenute nel file.

'''         
    
class XLSXReader(FormatReader):

    def size_extractor(self):
        
        file_size=super().size_extractor()
        
        return file_size
  
    def time_extractor(self):
        
        file_time=super().size_extractor()
        
        return file_time 
    
    def get_file_content(self):
        
         stringhe=' ' # Inizializzo una stringa vuota
         
         excel_document = openpyxl.load_workbook(self.filename) # Apro file excel
        
         nomi_fogli=(excel_document.sheetnames) # Salvo il nome dei fogli in una lista
   
         for nome in nomi_fogli: # Scandisco tutti i fogli del file excel
          
          sheet = excel_document[nome]

          for row in sheet.iter_rows(): # Scandisco le righe dei fogli excel
    
            for cell in row: # Scandisco le celle per riga
     
              if (cell.value is None)==False: # Se il valore della cella non è None...
       
                  stringhe=stringhe+((cell.value).lower())+' ' # ... aggiungo il contenuto della cella (in minuscolo) alla stringa                   
                 
                  lista_stringhe_xlsx=re.findall(r"[\w']+", stringhe)  # Elimino la punteggiatura e divido le stringhe ad ogni spazio
        
         return lista_stringhe_xlsx # Restituisco la lista delle parole contenute nel file

'''

Creo classe derivata per la lettura dei file pdf. Questa classe prende in ingresso il file pdf che sto analizzando e
restituisce la lista delle parole contenute nel file.

'''     
    

class PDFReader(FormatReader):
   
    def size_extractor(self):
        
        file_size=super().size_extractor()
        
        return file_size
  
    def time_extractor(self):
        
        file_time=super().size_extractor()
        
        return file_time 
   
    def get_file_content(self):
        
        with pdfplumber.open (self.filename) as pdf: # Apro il file pdf
           
           pages = pdf.pages  # Salvo le pagine del file pdf in una lista
           
           stringhe_pdf=' ' # Inizializzo una stringa vuota
           
           for page in pages: # Scandisco ogni pagina contenuta nella lista di pagine
                 
                 # Aggiungo il contenuto della pagina (in minuscolo) alla stringa
                 
                 stringhe_pdf=stringhe_pdf+ (page.extract_text ()).lower() + ' '  
           
           lista_stringhe_pdf=re.findall(r"[\w']+", stringhe_pdf)  # Elimino la punteggiatura e divido le stringhe ad ogni spazio
           
           
        return lista_stringhe_pdf # Restituisco la lista delle parole contenute nel file

'''
Creo la classe 'InputInterpreter' che restituisce, partendo dal file di input,
'lista_classi', ovvero la lista contenente tutte le condizioni da verificare. 
In particolare tale lista contiene la chiamata alle classi che verificano le singole condizioni.
Per ogni specifica condizione viene quindi creato un oggetto diverso.

'''


class InputInterpreter():
    
    def __init__(self,condition_list):
        
        self.condition_list=condition_list
       # self.image_objects=image_objects
        
    def extractor(self,image_objects):
        
        #inizializzo la lista delle classi da chiamare per verificare le condizioni
        lista_classi=[] 
        #Estraggo la lista delle differenti tipologie di condizione
        lista_condizioni= list(self.condition_list.keys())

        if 'size'  in lista_condizioni:
            
            '''
            Appendo a 'lista_classi' le diverse chiamate alla classe
            che gestisce la condizione sulla 'size' (dimensione del file). 
            Creo in questo modo due oggetti diversi, uno per soddisfare la
            dimensione minima e uno per la dimensione massima.
            '''
            min_value=self.condition_list['size']['min']
            max_value=self.condition_list['size']['max']
            lista_classi.append(SizeChecker(self.condition_list,min_value,max_value))
        
        if 'time'  in lista_condizioni:
            
            #Appendo a 'lista_classi' le diverse chiamate alla classe
            #che gestisce la condizione sul 'time' (data di creazione del file).
            
            min_value=self.condition_list['time']['min']
            max_value=self.condition_list['time']['max']
            lista_classi.append(TimeChecker(self.condition_list,min_value,max_value))  
        
        if 'wordlist' in lista_condizioni:
            
            '''
            Appendo a 'lista_classi' le diverse chiamate alla classe
            che gestisce la condizione sulla 'wordlist' (lista delle parole cercate).
            Creo tanti oggetti della classe 'OccurrenceChecker' quante sono le coppie parola-occorrenza cercate.
            '''
            for parola in self.condition_list['wordlist'].keys():
               lista_classi.append(OccurrenceChecker(parola,self.condition_list['wordlist'][parola]))
        
        if 'objectlist' in lista_condizioni:
            
            detector = ObjectDetection()
            
            model_path = "./models/yolo-tiny.h5"
            detector.setModelTypeAsTinyYOLOv3()
            detector.setModelPath(model_path)
            detector.loadModel()
            #Appendo a 'lista_classi' le diverse chiamate alla classe
            #che gestisce la condizione sulla 'objectlist' (lista degli oggetti cercati). 
            
            for obj in self.condition_list['objectlist'].keys():
                
                #controllo se l'oggetto è tra quelli riconoscibili dall'algoritmo
                    
                if obj in image_objects:
            
                    lista_classi.append(ImageChecker(self.condition_list,obj,self.condition_list['objectlist'][obj],detector))
                    
                 
        return lista_classi  #Restituisco la lista contenente la chiamata alle classi per ogni specifica condizione







################################ MAIN ########################################

# Apertura file input

            
with open(args.input_data) as json_file:
   
    condition_list=json.load(json_file)

image_objects = open(args.input_data_2, "r").read()    
    


lista_path=[] # Inizializzo lista che conterrà i path dei file trovati

for path in condition_list['dirlist']: # Scandisco tutti i path contenuti nel file input
 
  for root,dirs, files in os.walk(path, topdown=False): #Scandisco sottocartelle (dirs) e files
   
    for name in files: # Scandisco i file presenti nelle sottocartelle
       
       if name!= ".DS_Store":
        
             lista_path.append(os.path.join(root, name)) # Aggiungo il path del file 'name' alla lista dei path dei file trovati    
    

#Chiamando la classe 'InputInterpreter' estraggo la lista
#contenente le chiamate alle classi per ogni specifica condizione.  
    
lista_classi=InputInterpreter(condition_list).extractor(image_objects)   

# Inizializzo la lista che conterrà i file che soddisfano tutte le condizioni

searched_file=[]

# Scandisco i path dei file relativi alla porzione di filesystem su cui verificare le condizioni
for file in lista_path:     
    
    '''
    Il metodo 'create_instance' della classe 'Condition' rimanda, indistintamente 
    dal formato del file, alla classe 'PathChecker' che verifica che il formato sia 
    tra quelli cercati. In tal caso restituisce 'True', altrimenti 'failure'.
    Se il valore restituito è 'failure' passa direttamente ad analizzare il file successivo.
    '''
    
    condition=PathChecker(file,condition_list).function()  

    if condition !=False:
        
        '''
        Il metodo 'create_instance' della classe 'FormatReader' rimanda, in base all'estensione
        del file, alla classe derivata in grado di estrarne il contenuto, la size e la data di creazione.
        '''
        file_object=FormatReader.create_instance(file)
        
        '''
        Eseguo il ciclo su tutte le condizioni, chiamando il metodo 'checker' della classe
        'Condition', implementato diversamente per ogni tipologia di condizione. Il metodo checker 
        restituisce 'True' se la condizione è verificata e 'False' in caso contrario.
        '''
        
        for condizione in lista_classi:
            
            is_verified_condition= condizione.checker(file_object)
              
            #Se la condizione non è verificata esco dal ciclo passando al file successivo
            if  is_verified_condition== False:  
                 
                break 
            
        #Se verifica tutte le condizioni lo appendo nel file di output 'searched_file'
        if  is_verified_condition== True:       
              
              searched_file.append(file)
              
            
# Salvo su un file in formato txt la lista dei file che rispettano le condizioni

with open(args.out_data, "w") as output:
    
    output.write(str(searched_file))  